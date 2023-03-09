# Import Section

import pygame
import time
import openpyxl
import random

# Defining Some Variables for GAME
  
user,passwd = '',''
USER_IN = ''
X,Y = 0,0
current_time = ['']
start_time = ''
Level = 'Easy'
t = 0
l_y1 = False
l_y2 = False
l_y3 = False
yx,yy = 0,0

# Color defination

RED = (255, 0, 0)
GREEN = (0, 255, 0)
BLUE = (0, 0, 255)
LIGHT_BLUE   = (0,0,150)
YELLOW = (255, 255, 0)
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
PURPLE = (128, 0, 128)
ORANGE = (255, 165 ,0)
GREY = (128, 128, 128)
TURQUOISE = (64, 224, 208)

######### LOADING GAME IMAGES ############

icon = pygame.image.load('./image/icon.png')
login_back = pygame.image.load("./image/Login_back1.jpg")
setting = pygame.image.load("./image/setting.png")
back = pygame.image.load("./image/back.png")
hero_yellow = pygame.image.load("./image/pixel_ship_yellow.png")
enemy_red = pygame.image.load('./image/pixel_ship_red_small.png')
enemy_blue = pygame.image.load('./image/pixel_ship_blue_small.png')
enemy_green = pygame.image.load('./image/pixel_ship_green_small.png')
laser_yellow = pygame.image.load('./image/pixel_laser_yellow.png')
laser_red = pygame.image.load('./image/pixel_laser_red.png')
laser_blue = pygame.image.load('./image/pixel_laser_blue.png')
laser_green = pygame.image.load('./image/pixel_laser_green.png')

# Creating Class

class enemy(object):
    def __init__(self,width,height):
        self.ex1 = random.choice([i for i in range(int(width*0.5),int(width*17.5))])
        self.ey1 = height+5
        self.width = width
        self.height = height
        self.enemy_red = pygame.image.load('./image/pixel_ship_red_small.png')
        self.enemy_blue = pygame.image.load('./image/pixel_ship_blue_small.png')
        self.enemy_green = pygame.image.load('./image/pixel_ship_green_small.png')
        self.laser_red = pygame.image.load('./image/pixel_laser_red.png')
        self.laser_blue = pygame.image.load('./image/pixel_laser_blue.png')
        self.laser_green = pygame.image.load('./image/pixel_laser_green.png')
        self.redefine()
        self.life_line = 100

    def redefine(self):    
        self.lx = self.ex1
        self.ly = self.ey1

    def surface(self,color):
        self.color = color
        if color == 'b':
            return self.enemy_blue
        if color == 'g':
            return self.enemy_green
        if color == 'r':
            return self.enemy_red
    def cordinates(self,cordi):
        self.hx = cordi[0]
        self.hy = cordi[1]
        if self.hx<self.ex1:
            self.ex1 -= self.width*0.02
        if self.hx>self.ex1:
            self.ex1 += self.width*0.02
        return (int(self.ex1),int(self.ey1))
    def laser(self):
        if self.color == 'b':
            return self.laser_blue
        if self.color == 'g':
            return self.laser_green
        if self.color == 'r':
            return self.laser_red
    def l_cordinates(self):
        if self.ly<self.height*20:
            self.ly += self.height*0.2
        if self.ly >= self.height*20:
            self.redefine()
        return (int(self.lx-40),int(self.ly))
    def life(self,laser_list):
        self.yx1=laser_list[0]
        self.yy1=laser_list[1]
        self.yx2=laser_list[2]
        self.yy2=laser_list[3]
        self.yx3=laser_list[4]
        self.yy3=laser_list[5]
        if self.color == 'b':
            if ((self.ex1-20<self.yx1<self.ex1) or (self.ex1-20<self.yx2<self.ex1) or (self.ex1-20<self.yx3<self.ex1)) and ((self.ey1+30>=self.yy1) or (self.ey1+30>=self.yy2) or (self.ey1+30>=self.yy3)):
                self.life_line -= 25
                return [True,(int(self.ex1),int(self.ey1),120,24),(int(self.ex1+10),int(self.ey1+2),int(self.life_line),20)]
    def kill(self):
        if self.life_line<=0:
            return True
    def remove(self):
        self.ex1 = self.width*21+1
    def remake(self):
        self.ex1 = random.choice([i for i in range(int(width*0.5),int(width*17.5))])
        self.life_line = 100
# Creating Display (grid system)

WIDTH,HEIGHT = 1200,700
width = WIDTH//20   
height = HEIGHT//20
WIN = pygame.display.set_mode((WIDTH,HEIGHT))
pygame.display.set_caption('Space Invader')
    # Setting Icon
pygame.display.set_icon(icon)

# creating constant for images

hx,hy = width*9,height*17
blue = enemy(width,height)
######## Creating Login and High Score Database File ########

    # user passwd DICT
user_pass = {'Ankit':'ankit231',
             'Ankita':'ankita1802'}
    # High Score DICT
user_score = {'Ankit':0,
              'Ankita':0}
wb = openpyxl.load_workbook('High_Score.xlsx')
ws = wb.active
for i in range(len(list(user_pass.keys()))):
    ws.cell(1,i+1,list(user_pass.keys())[i])
    if user_score[list(user_pass.keys())[i]] <= ws.cell(2,i+1).value:
        user_score[list(user_pass.keys())[i]] = ws.cell(2,i+1).value

# SOME IMPORTANT THINGS TO WORK

    # Initialising Pygame
pygame.init()
    # seting repeat for key
pygame.key.set_repeat(100)

######## CREATING MAIN LOOP #########

    # Constant for MAIN LOOP
RUN = True
LOGINPAGE = True
GAME = False
SETTING = False
    # Main Loop STARTs
while RUN:
    WIN.fill(BLACK)
        # Geting Events
    for event in pygame.event.get():
        if event.type == pygame.QUIT :
            i = list(user_pass.keys()).index(user)
            if user_score[user] >= ws.cell(2,i+1).value:
                ws.cell(2,i+1,user_score[user])
            wb.save('High_Score.xlsx')
            RUN = False
        if event.type == pygame.MOUSEBUTTONDOWN:
            X,Y = pygame.mouse.get_pos()
            # print(f"mouse = {X,Y}")
            if GAME:
                if event.button == 1:
                    if not l_y1:
                        yx1,yy1 = hx,hy
                        l_y1 = True
                    elif not l_y2:
                        yx2,yy2 = hx,hy
                        l_y2 = True
                    elif not l_y3:
                        yx3,yy3 = hx,hy
                        l_y3 = True
                    # fire = True
        if event.type == pygame.KEYDOWN:
            USER_IN = event.unicode
            if width*4<X<width*10 and height*6<Y<height*6.7:
                user = user+USER_IN
                if event.key == 8:
                    user = user[:-2]
            if width*4<X<width*10 and height*8<Y<height*8.7:
                passwd =passwd+USER_IN
                if event.key == 8:
                   passwd =passwd[:-2]
            if GAME:
                if event.key == 273 or event.key == 119:
                    if hy > height*13:
                        hy = hy - int(height*0.2)
                if event.key == 274 or event.key == 115:
                    if hy < height*18:
                        hy = hy + int(height*0.2)
                if event.key == 275 or event.key == 100:
                    if hx < int(width*17.5):
                        hx = hx + int(width*0.2)
                if event.key == 276 or event.key == 97:
                    if hx > int(width*0.5):
                        hx = hx - int(width*0.2)
        # Creting Login Page
    if LOGINPAGE:
        WIN.blit(login_back,(0,0))
        WIN.blit(pygame.font.SysFont('Algerian',30,0,1).render('Login Page',0,ORANGE),(width*8,height*2))
        WIN.blit(pygame.font.SysFont('Comicstan',32,0,1).render('User Name',0,ORANGE),(width,height*6))
        pygame.draw.rect(WIN,WHITE,(width*4,height*6,width*6,int(height*0.7)))
        if width*4<X<width*10 and height*6<Y<int(height*6.7):
            pygame.draw.rect(WIN,RED,(int(width*3.97),int(height*5.95),int(width*6.15),int(height*0.8)))
            pygame.draw.rect(WIN,WHITE,(width*4,height*6,width*6,int(height*0.7)))
        WIN.blit(pygame.font.SysFont('Arial',26,0,1).render(user,1,BLACK),(int(width*4.02),int(height*5.98)))
        WIN.blit(pygame.font.SysFont('Comicstan',32,0,1).render('Password',0,ORANGE),(width,height*8))
        pygame.draw.rect(WIN,WHITE,(width*4,height*8,width*6,int(height*0.7)))
        if width*4<X<width*10 and height*8<Y<int(height*8.7):
            pygame.draw.rect(WIN,RED,(int(width*3.97),int(height*7.95),int(width*6.15),int(height*0.8)))
            pygame.draw.rect(WIN,WHITE,(width*4,height*8,width*6,int(height*0.7)))
        WIN.blit(pygame.font.SysFont('Arial',26,0,1).render(passwd,1,BLACK),(int(width*4.02),int(height*7.98)))
        WIN.blit(pygame.font.SysFont('Comicstan',40,0,1).render('NEXT',0,GREEN,BLUE),(width*7,height*10))
        if width*7<X<width*9 and height*10<Y<height*11 :
            for x,y in user_pass.items():
                if user == x and passwd == y:
                    LOGINPAGE=False
                    GAME = True
                elif user == ''and passwd == '':
                    WIN.blit(pygame.font.SysFont('Comicstan',35,0,1).render('Type user Name and Passwd',0,RED),(width*4,height*14))
                elif user == x and passwd != y:
                    WIN.blit(pygame.font.SysFont('Comicstan',35,0,1).render('Wrong Password',0,RED),(width*4,height*14))
        WIN.blit(setting,(width*18,height*18))
        if width*18<X<int(width*19.3) and height*18<Y<int(height*19.5) :
            SETTING = True
    if SETTING:
        WIN.fill(GREY)
        WIN.blit(back,(width,height))
        if int(width*1.25)<X<width*2 and height<Y<height*3:
            SETTING = False
            LOGINPAGE = True
        if Level == 'Easy':
            pygame.draw.rect(WIN,LIGHT_BLUE,(int(width*4.8),int(height*5.9),int(width*3.6),int(height*1.5)))
        if Level == 'Medium':
            pygame.draw.rect(WIN,LIGHT_BLUE,(int(width*4.8),int(height*7.9),int(width*3.6),int(height*1.5)))
        if Level == 'Hard':
            pygame.draw.rect(WIN,LIGHT_BLUE,(int(width*4.8),int(height*9.9),int(width*3.6),int(height*1.5)))
        WIN.blit(pygame.font.SysFont('Comicstan',50,1,1).render('Select Game Level',0,ORANGE),(width*4,height*4))
        WIN.blit(pygame.font.SysFont('Arial',40,0,1).render('Easy',1,GREEN),(width*5,height*6))
        WIN.blit(pygame.font.SysFont('Arial',40,0,1).render('Medium',1,YELLOW),(width*5,height*8))
        WIN.blit(pygame.font.SysFont('Arial',40,0,1).render('Hard',1,RED),(width*5,height*10))
        if width*5<X<width*7 and height*6<Y<int(height*7.1):
            Level = 'Easy'
        if width*5<X<int(width*7.5) and height*8<Y<int(height*9.1):
            Level = 'Medium'
        if width*5<X<width*7 and height*10<Y<int(height*11.1):
            Level = 'Hard'
    if GAME:
        WIN.blit(login_back,(0,0))
        if t == 0:
            WIN.blit(pygame.font.SysFont('Algerian',50,0,1).render("Gane start",0,BLACK,ORANGE),(width*6,height*10))
            pygame.display.update()
            time.sleep(.5)
            t += 1
        current_time[0] = time.strftime("%H:%M:%S",time.localtime()).split(':')
        if (start_time == '') and (current_time[0] != ''):
            start_time=current_time[0]
        Current_score = (int(current_time[0][0])*60*60)+(int(current_time[0][1])*60)+(int(current_time[0][2]))-((int(start_time[0])*60*60)+(int(start_time[1])*60)+(int(start_time[2]))) 
        if Current_score>=user_score[user]:
            user_score[user] = Current_score
        pygame.draw.rect(WIN,ORANGE,(0,0,WIDTH,height))
        WIN.blit(pygame.font.SysFont('Algerian',20,0,1).render("User",0,BLACK),(int(width*0.8),int(height*0.1)))
        WIN.blit(pygame.font.SysFont('Arial',26,0,1).render(f"{user}",1,BLACK),(width*2,int(height*0.1)))
        WIN.blit(pygame.font.SysFont('Algerian',20,0,1).render("Level",0,BLACK),(width*4,int(height*0.1)))
        WIN.blit(pygame.font.SysFont('Arial',26,0,1).render(f"{Level}",1,BLACK),(int(width*5.8),int(height*0.1)))
        WIN.blit(pygame.font.SysFont('Algerian',20,0,1).render("High Score",0,BLACK),(width*9,int(height*0.1)))
        WIN.blit(pygame.font.SysFont('Arial',20,0,1).render(f"{user_score[user]}",0,BLACK),(width*12,int(height*0.1)))
        WIN.blit(pygame.font.SysFont('Algerian',20,0,1).render("Score",0,BLACK),(width*16,int(height*0.1)))
        WIN.blit(pygame.font.SysFont('Arial',20,0,1).render(f"{Current_score}",0,BLACK),(int(width*17.8),int(height*0.1)))
        pygame.draw.rect(WIN,BLACK,(0,height,3,height*19))
        pygame.draw.rect(WIN,BLACK,(0,HEIGHT-3,WIDTH,3))
        pygame.draw.rect(WIN,BLACK,(WIDTH-3,height,3,height*19))
        WIN.blit(hero_yellow,(hx,hy))
        # print(f"hx,hy = {(hx,hy)}")
        try:
            WIN.blit(blue.surface('b'),blue.cordinates([hx,hy]))
            # print(f"ex = {blue.cordinates([hx,hy])}")
            WIN.blit(blue.laser(),blue.l_cordinates())
        except:
            pass
        try:
            life_blue = blue.life([yx1,yy1,yx2,yy2,yx3,yy3])
            # print(f"cordinate = {[yx1,yy1,yx2,yy2,yx3,yy3]}")
        except:
            pass
        if l_y1:
            WIN.blit(laser_yellow,(int(yx1),int(yy1)))
            yy1 -= height*0.2
            if yy1 < height:
                l_y1 = False
                yx1,yy1 = WIDTH,HEIGHT
        if l_y2:
            WIN.blit(laser_yellow,(int(yx2),int(yy2)))
            yy2 -= height*0.2
            if yy2 < height:
                l_y2 = False
                yx2,yy2 = WIDTH,HEIGHT
        if l_y3:
            WIN.blit(laser_yellow,(int(yx3),int(yy3)))
            yy3 -= height*0.2
            if yy3 < height:
                l_y3 = False
                yx3,yy3 = WIDTH,HEIGHT
        try:
            if life_blue[0]:
                pygame.draw.rect(WIN,WHITE,life_blue[1])
                pygame.draw.rect(WIN,BLUE,life_blue[2])
        except:
            pass
        if blue.kill():
            blue.remove()
            time.sleep(0.10)
            blue.remake()
    pygame.display.update()